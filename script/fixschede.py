from openpyxl import Workbook
from openpyxl import load_workbook
from pathlib import Path
import sys

app_dir = Path(sys.argv[0]).parent

dir_schede = Path(sys.argv[1] if len(sys.argv) >= 2 else 'W:/schede')

scheda_pulita = dir_schede.joinpath('--- BE Pulito.xlsx')

for f in dir_schede.rglob('*.xlsx'):
    if f == scheda_pulita: continue

    print (f)
    wb_new = load_workbook(scheda_pulita)

    wb = load_workbook(f)
    for i, row in enumerate(wb['In Letture']['B3:Q39'], 3):
        for j, cell in enumerate(row, 2):
            c_new = wb_new['In Letture'].cell(row = i, column = j)
            c_new.value = cell.value
            c_new.number_format = cell.number_format
    
    for i, row in enumerate(wb['In Dati']['B3:Z38'], 3):
        for j, cell in enumerate(row, 2):
            c_new = wb_new['In Dati'].cell(row = i, column = j)
            c_new.value = cell.value
            c_new.number_format = cell.number_format

    wb.close()
    wb_new.save(f)