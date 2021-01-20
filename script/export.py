from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side
import sys
import json
import time
import os

class TableValue:
    def __init__(self, title, value, index=0, type='str', number_format='', column_width=None, obligatory=False):
        self.title = title
        self.value = value
        self.index = index
        self.type = type
        self.number_format = number_format
        self.column_width = column_width
        self.obligatory = obligatory

table_values = [
    TableValue('File',                  'filename'),
    TableValue('POD',                   'codice_pod', column_width=16,            obligatory=True),
    TableValue('Mese',                  'mese_fattura',             type='month', obligatory=True),
    TableValue('Fornitore',             'fornitore',                              obligatory=True),
    TableValue('N. Fatt.',              'numero_fattura',                         obligatory=True, column_width=11),
    TableValue('Data Emissione',        'data_fattura',              type='date', obligatory=True),
    TableValue('Data scadenza',         'data_scadenza',             type='date'),
    TableValue('Costo Materia Energia', 'spesa_materia_energia',     type='euro', column_width=11),
    TableValue('Trasporto',             'trasporto_gestione',        type='euro', column_width=11),
    TableValue('Oneri',                 'oneri',                     type='euro', column_width=11),
    TableValue('Accise',                'accise',                    type='euro', column_width=11),
    TableValue('Iva',                   'iva',                       type='percentage', column_width=6),
    TableValue('Imponibile',            'imponibile',                type='euro', column_width=11),
    # TableValue('F1 Rilev.',             'energia_attiva_rilevata', index=0, type='int', column_width=8),
    # TableValue('F2 Rilev.',             'energia_attiva_rilevata', index=1, type='int', column_width=8),
    # TableValue('F3 Rilev.',             'energia_attiva_rilevata', index=2, type='int', column_width=8),
    TableValue('F1',                    'energia_attiva',   index=0, type='int', column_width=8),
    TableValue('F2',                    'energia_attiva',   index=1, type='int', column_width=8),
    TableValue('F3',                    'energia_attiva',   index=2, type='int', column_width=8),
    TableValue('P1',                    'potenza',          index=0, type='int', column_width=8),
    TableValue('P2',                    'potenza',          index=1, type='int', column_width=8),
    TableValue('P3',                    'potenza',          index=2, type='int', column_width=8),
    TableValue('R1',                    'energia_reattiva', index=0, type='int', column_width=8),
    TableValue('R2',                    'energia_reattiva', index=1, type='int', column_width=8),
    TableValue('R3',                    'energia_reattiva', index=2, type='int', column_width=8),
    TableValue('Oneri Ammin.',          'oneri_amministrativi',      type='euro'),
    TableValue('PCV',                   'pcv',                       type='euro'),
    TableValue('CTS',                   'cts',                       type='euro'),
    TableValue('<75%',                  'penale_reattiva_inf75',     type='euro', column_width=11),
    TableValue('>75%',                  'penale_reattiva_sup75',     type='euro', column_width=11),
    TableValue('PEF1',                  'prezzo_energia',   index=0, type='number', number_format='0.00000000', column_width=11),
    TableValue('PEF2',                  'prezzo_energia',   index=1, type='number', number_format='0.00000000', column_width=11),
    TableValue('PEF3',                  'prezzo_energia',   index=2, type='number', number_format='0.00000000', column_width=11),
    TableValue('Sbilanciamento',        'sbilanciamento',            type='number', number_format='0.00000000', column_width=11),
    TableValue('Disp. Var',             'disp_var',                  type='number', number_format='0.00000000', column_width=11)
]

old_pod = None

def export_file(input_file):
    out = []
    out_err = []

    def add_rows(json_data):
        filename = json_data['filename']

        if 'error' in json_data:
            out_err.append([filename, json_data['error']])
            return
        
        for v in json_data['values']:
            global old_pod
            row = []

            for obj in table_values:
                if obj.value == 'filename':
                    row.append({'value':filename, 'number_format': ''})
                else:
                    try:
                        value = v[obj.value][obj.index]
                        if obj.type == 'str':
                            row.append({'value': value, 'number_format': '@'})
                        elif obj.type == 'date':
                            row.append({'value': datetime.strptime(value, '%Y-%m-%d'), 'number_format': 'DD/MM/YY'})
                        elif obj.type == 'month':
                            row.append({'value': datetime.strptime(value, '%Y-%m'), 'number_format': 'MM/YYYY'})
                        elif obj.type == 'euro':
                            row.append({'value': float(value), 'number_format': '#,##0.00 \u20ac'})
                        elif obj.type == 'int':
                            row.append({'value': float(value), 'number_format': '0'})
                        elif obj.type == 'number':
                            row.append({'value': float(value), 'number_format': obj.number_format})
                        elif obj.type == 'percentage':
                            row.append({'value': float(value[:-1]) / 100, 'number_format': '0%'})
                        else:
                            row.append({'value': value, 'number_format': ''})
                    except (KeyError, IndexError, ValueError):
                        if obj.obligatory:
                            continue
                        else:
                            row.append({'value': '', 'number_format': ''})

            new_pod = v['codice_pod'] if 'codice_pod' in v else None
            out.append({'row':row,'conguaglio':'conguaglio' in json_data,'new_pod':new_pod != old_pod})
            old_pod = new_pod

    with open(input_file, 'r') as fin:
        for r in json.loads(fin.read()):
            add_rows(r)

    wb = Workbook()
    ws = wb.active

    ws.append([obj.title for obj in table_values])
    for i, obj in enumerate(table_values, 1):
        if obj.column_width != None:
            ws.column_dimensions[get_column_letter(i)].width = obj.column_width

    for i, row in enumerate(out, 2):
        for j, c in enumerate(row['row'], 1):
            cell = ws.cell(row=i, column=j)
            if row['new_pod']:
                cell.border = Border(top=Side(border_style='thin', color='000000'))
            if row['conguaglio']:
                cell.fill = PatternFill(patternType='solid', fgColor='ffff00')
            try:
                cell.number_format = c['number_format']
                cell.value = c['value']
            except (KeyError, IndexError, ValueError):
                pass

    for row in out_err:
        ws.append(row)

    ws.freeze_panes = ws['A2']

    wb.save(input_file.with_suffix('.xlsx'))

if __name__ == '__main__':
    if len(sys.argv) <= 1:
        print('Specificare il file di input')
        sys.exit(0)
    
    export_file(Path(sys.argv[1]))