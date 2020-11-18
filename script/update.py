from openpyxl import Workbook
from openpyxl import load_workbook
from pathlib import Path
import datetime
import sys
from shutil import copyfile

app_dir = Path(sys.argv[0]).parent

dir_out = Path(sys.argv[1]) if len(sys.argv) >= 2 else app_dir.joinpath('out')
dir_schede = Path(sys.argv[2] if len(sys.argv) >= 3 else 'W:/schede')

scheda_pulita = dir_schede.joinpath('--- BE Pulito.xlsx')

for f in dir_out.rglob('*.xlsx'):
    nome_cliente = f.stem

    try:
        wb = load_workbook(f)
        ws = wb.active
    except:
        continue

    old_filename = ''

    indices = {c.value.strip():i for i,c in enumerate(tuple(ws.rows)[0]) if isinstance(c.value, str)}

    for row in tuple(ws.rows)[1:]:
        codice_pod = row[indices['POD']].value
        mese_fattura = row[indices['Mese']].value
        num_fattura = row[indices['N. Fatt.']].value
        if num_fattura == None or codice_pod == None:
            continue
        if mese_fattura != None and not isinstance(mese_fattura, datetime.datetime):
            mese_fattura = datetime.datetime.strptime(mese_fattura, '%Y-%m')


        filename = dir_schede.joinpath('BE {0} {1}.xlsx'.format(nome_cliente, codice_pod[-3:]))
        if not filename.exists():
            filename = dir_schede.joinpath('BE {0} {1}.xlsx'.format(nome_cliente, codice_pod[-4:]))
        
        if filename != old_filename:
            if filename.exists():
                newfile = False
            else:
                copyfile(scheda_pulita, filename)
                newfile = True
            
            if old_filename != '':
                wb_scheda.save(old_filename)
            try:
                wb_scheda = load_workbook(filename)
                ws_scheda = wb_scheda['In Dati']
            except FileNotFoundError:
                print(filename, 'Non trovato')
                continue
            old_filename = filename
            
        indices_scheda = {c.value.strip():i for i,c in enumerate(tuple(ws_scheda.rows)[1]) if isinstance(c.value, str)}

        found = False
        isnew = False
        for row_scheda in tuple(ws_scheda.rows)[2:]:
            mese_scheda = row_scheda[indices_scheda['Mese fattura']].value
            if not isinstance(mese_scheda, datetime.datetime):
                mese_scheda = datetime.datetime.strptime(mese_scheda, '%Y-%m')
            if mese_fattura != None and mese_fattura != '' and mese_scheda == mese_fattura:
                num_fattura_scheda = row_scheda[indices_scheda['N. Fatt.']].value
                if num_fattura_scheda == '' or num_fattura_scheda == None:
                    isnew = True
                elif num_fattura == num_fattura_scheda:
                    found = True
                else:
                    continue
                for cell_out, cell_scheda in zip(row[indices['Fornitore']:indices['PEF3']+1], row_scheda[indices_scheda['Fornitore']:indices_scheda['PEF3']+1]):
                    if isinstance(cell_out.value,str) and cell_out.value[-1:] == '%':
                        cell_out.number_format = '0%'
                        cell_out.value = float(cell_out.value[:-1]) / 100
                    elif isinstance(cell_out.value,datetime.datetime):
                        cell_out.number_format = 'DD/MM/YY'
                    cell_scheda.number_format = cell_out.number_format
                    cell_scheda.value = cell_out.value
        print('Added\t' if isnew else 'Update\t' if found else 'Skip\t',
            nome_cliente, codice_pod, num_fattura,
            mese_fattura.date() if isinstance(mese_fattura, datetime.datetime) else '-')
    
    wb_scheda.save(filename)
